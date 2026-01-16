import sys
import os
import shutil
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, DEFAULT_FONT, Color, PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList

auto_open_flg = True

LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
def n2c(n:int):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while n:
        n, rem = divmod(n-1, 26)
        result[:0] = LETTERS[rem]
    return ''.join(result)
def c2n(name:str):
    """ Converts an Excel-style column name (e.g., 'A', 'B', 'AA') to its 1-based numerical index. """
    index = 0
    for char in name.upper():
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index # Adjust for 0-based indexing

def next_col(col) -> str:
    return n2c(c2n(col) + 1)

# styles
RedFill = PatternFill(start_color='EE1111', end_color='EE1111', bgColor='EE1111', fill_type='solid')
norm_side = Side(border_style='thin', color='000000')

def normal_border_style(directs):
    kwargs = {}
    if 'h' in directs:
        kwargs['left'] = norm_side
    if 'l' in directs:
        kwargs['right'] = norm_side
    if 'k' in directs:
        kwargs['top'] = norm_side
    if 'j' in directs:
        kwargs['bottom'] = norm_side
    return Border(**kwargs)
nbs = normal_border_style

def set_text(cell):
    cell.number_format = '@'
st = set_text

def set_link(workbook, worksheet, pos, destsheet, destpos, back_str='', back_linkpos = ''):
    if worksheet[pos].value is None:
        return
    hyperlink = Font(underline='single', color='0563C1')
    worksheet[pos].font = hyperlink
    if destsheet:
        worksheet[pos].hyperlink = f'#{destsheet}!{destpos}'
    else:
        worksheet[pos].hyperlink = f'#{worksheet.title}!{destpos}'
    if back_str: # set back link
        if back_linkpos == '':
            back_linkpos = destpos
        _ws = workbook[destsheet] if destsheet else worksheet
        _ws[back_linkpos] = back_str
        _ws[back_linkpos].font = hyperlink
        _ws[back_linkpos].hyperlink = f'#{worksheet.title}!{pos}'

def set_conditional_format(worksheet, cell_range, formula, font=DEFAULT_FONT, border=None, fill=None):
    kwargs = {'font' : font}
    if border: kwargs['border'] = border
    if fill: kwargs['fill'] = fill

    fr = FormulaRule(formula=[formula], stopIfTrue=False, **kwargs)
    worksheet.conditional_formatting.add(cell_range, fr)
scf = set_conditional_format

def clear_conditional_format(worksheet):
    '''clear conditional formatting for whole worksheet'''
    worksheet.conditional_formatting = ConditionalFormattingList()

def reflect_formula(formula:str, src:str, dst:str):
    '''generate new excel formula string based on original formula and cell'''
    return formula.upper().replace(src.upper(), dst)
rf = reflect_formula # short cut

def limit_input_formula(pos, vals:list):
    _limit_formula = ','.join(f'not(exact({pos}, "{v}"))' for v in vals)
    return f'AND(NOT(ISBLANK({pos})), {_limit_formula})'.upper()

def not_formula(formula):
    if formula[0] == '=':
        return f'=not{formula[1:]}'
    return f'=not{formula}'

# ############################## ↓↓↓ script-here ↓↓↓ ##############################

def process_sheet(wb, ws):
    ##################### examples #####################

    # row = i + 2
    # col = 'F'
    # ws[f'{col}{row}'] = f'content' # set content
    # set_link(wb, ws, f'{col}{row}', f'No{i}', 'A1') # set link

    ##################### examples #####################
    pass



def process(fp):
    wb = openpyxl.load_workbook(fp) # load workbook
    # print(wb.sheetnames)

    ##################### examples #####################

    # wb.create_sheet(f'No{i}') # create sheet

    ##################### examples #####################

    ws = wb['Sheet1'] # active worksheet
    process_sheet(wb, ws)





    wb.save(fp) # save workbook
    wb.close()

# ############################## ↑↑↑ script here ↑↑↑ ##############################

def to_win_path(wsl_path):
    path = wsl_path[5:]
    return f'{path[0]}:{path[1:]}'

def backup(fp):
    timestamp = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    dst = f'{root}-{timestamp}{extension}'
    shutil.copyfile(fp, dst)
    print(f'backup to {dst}')

if __name__ == '__main__':
    if len(sys.argv) < 2: print('No file given... exit'); exit(0)
    fp = sys.argv[1]
    root, extension = os.path.splitext(fp)
    if extension != '.xlsx': print('Not excel file... exit'); exit(0)
    # back up first
    backup(fp)

    print('Start processing...')
    process(fp)
    print('Processed')

    if auto_open_flg:
        print('Opening...')
        os.system(f'wslview {to_win_path(fp)}')
        print('Done')
